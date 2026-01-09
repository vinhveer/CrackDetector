import argparse
import csv
import datetime as dt
import json
import os
import time
from pathlib import Path
from typing import Any

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from src.pipeline.pipeline import CrackDetectionPipeline


IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"}


STAGE_ORDER: list[tuple[str, str]] = [
    ("input", "input"),
    ("preprocessed", "preprocess"),
    ("dino_boxes_overlay", "dino_boxes"),
    ("sam_raw_mask_viz", "sam_raw_mask"),
    ("geometry_input_viz", "geometry_input"),
    ("geometry_kept_mask_viz", "geometry_kept"),
    ("final_mask", "final_mask"),
    ("final_overlay", "final_overlay"),
]


def _safe_name(p: Path) -> str:
    return p.stem.replace(" ", "_").replace(os.sep, "_")


def _to_uint8_image(img: np.ndarray) -> np.ndarray:
    if img.dtype == np.bool_:
        return (img.astype(np.uint8) * 255)

    if img.dtype == np.uint8:
        return img

    arr = img

    if np.issubdtype(arr.dtype, np.floating):
        arr = np.nan_to_num(arr, nan=0.0, posinf=0.0, neginf=0.0)
        mn = float(np.min(arr)) if arr.size else 0.0
        mx = float(np.max(arr)) if arr.size else 0.0
        if mx <= 255.0 and mn >= 0.0:
            return np.clip(arr, 0.0, 255.0).astype(np.uint8)
        if mx == mn:
            return np.zeros(arr.shape, dtype=np.uint8)
        norm = cv2.normalize(arr, None, 0, 255, cv2.NORM_MINMAX)
        return norm.astype(np.uint8)

    if np.issubdtype(arr.dtype, np.integer):
        mn = int(arr.min()) if arr.size else 0
        mx = int(arr.max()) if arr.size else 0
        if 0 <= mn and mx <= 255:
            return arr.astype(np.uint8)
        if mx == mn:
            return np.zeros(arr.shape, dtype=np.uint8)
        norm = cv2.normalize(arr.astype(np.float32), None, 0, 255, cv2.NORM_MINMAX)
        return norm.astype(np.uint8)

    return arr.astype(np.uint8)


def _save_step_image(step_dir: Path, base_name: str, img: np.ndarray) -> Path:
    out = img
    if out.ndim == 3 and out.shape[2] == 1:
        out = out[:, :, 0]

    out = _to_uint8_image(out)

    ext = ".png" if out.ndim == 2 else ".jpg"
    out_path = step_dir / f"{base_name}{ext}"
    cv2.imwrite(str(out_path), out)
    return out_path


def _image_stats(img: np.ndarray) -> dict[str, Any]:
    out: dict[str, Any] = {
        "shape": json.dumps(list(img.shape)),
        "dtype": str(img.dtype),
    }
    try:
        out["nonzero"] = int(np.count_nonzero(img))
    except Exception:
        out["nonzero"] = None

    if img.size == 0:
        out["min"] = None
        out["max"] = None
        out["sum"] = None
        return out

    try:
        arr = img
        if np.issubdtype(arr.dtype, np.floating):
            arr = np.nan_to_num(arr, nan=0.0, posinf=0.0, neginf=0.0)
        out["min"] = float(np.min(arr))
        out["max"] = float(np.max(arr))
        out["sum"] = float(np.sum(arr))
    except Exception:
        out["min"] = None
        out["max"] = None
        out["sum"] = None
    return out


def _build_stage_list(images: dict[str, Any]) -> list[tuple[str, str]]:
    # Return list of (key, display_name) with deterministic order.
    present = set((images or {}).keys())
    stages: list[tuple[str, str]] = [(k, n) for (k, n) in STAGE_ORDER if k in present]

    # Add remaining keys not in STAGE_ORDER, sorted for stability.
    used = {k for k, _ in stages}
    extras = sorted([k for k in present if k not in used])
    for k in extras:
        stages.append((k, k))
    return stages


def _write_metrics_csv(path: Path, row: dict[str, Any]) -> None:
    # store a single-row CSV for easy import
    keys = sorted(row.keys())
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        w.writerow({k: row.get(k) for k in keys})


def _write_aggregated_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return

    keys_set: set[str] = set()
    for r in rows:
        keys_set.update(r.keys())

    preferred = ["image", "image_path", "status", "error"]
    fieldnames: list[str] = []
    for k in preferred:
        if k in keys_set:
            fieldnames.append(k)
            keys_set.remove(k)
    fieldnames.extend(sorted(keys_set))

    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k) for k in fieldnames})


def _write_regions_csv(path: Path, regions: list[Any]) -> None:
    if not regions:
        return

    rows: list[dict[str, Any]] = []
    for reg in regions:
        row: dict[str, Any] = {}
        row["region_id"] = int(getattr(reg, "region_id", 0))
        row["kept"] = bool(getattr(reg, "kept", False))
        row["dropped_reason"] = getattr(reg, "dropped_reason", None)

        bbox = getattr(reg, "bbox", None)
        if bbox is not None:
            try:
                row["bbox"] = json.dumps(list(bbox))
            except Exception:
                row["bbox"] = str(bbox)

        geom = getattr(reg, "geometry", None)
        if geom is not None:
            for k in ("area", "skeleton_length", "endpoint_count", "width_mean", "width_var", "length_area_ratio", "curvature_var"):
                if hasattr(geom, k):
                    row[k] = getattr(geom, k)

        scores = getattr(reg, "scores", None)
        if scores is not None:
            for k in ("dino_conf", "final_conf"):
                if hasattr(scores, k):
                    row[f"score_{k}"] = getattr(scores, k)

        meta = getattr(reg, "meta", None)
        if meta is not None:
            for k in ("prompt_name", "variant_name", "damage_type"):
                if hasattr(meta, k):
                    row[f"meta_{k}"] = getattr(meta, k)

        mask = getattr(reg, "mask", None)
        if isinstance(mask, np.ndarray):
            try:
                row["mask_sum"] = int((mask > 0).sum())
            except Exception:
                row["mask_sum"] = None

        rows.append(row)

    keys_set: set[str] = set()
    for r in rows:
        keys_set.update(r.keys())
    fieldnames = ["region_id", "kept", "dropped_reason"]
    for k in sorted(keys_set):
        if k not in fieldnames:
            fieldnames.append(k)

    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k) for k in fieldnames})


def _write_stages_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    keys_set: set[str] = set()
    for r in rows:
        keys_set.update(r.keys())
    preferred = ["stage_index", "stage_key", "stage_name", "filename", "shape", "dtype", "min", "max", "sum", "nonzero"]
    fieldnames: list[str] = []
    for k in preferred:
        if k in keys_set:
            fieldnames.append(k)
            keys_set.remove(k)
    fieldnames.extend(sorted(keys_set))
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k) for k in fieldnames})


def iter_images(input_dir: Path, recursive: bool) -> list[Path]:
    if not input_dir.exists():
        return []

    if recursive:
        paths = [p for p in input_dir.rglob("*") if p.is_file()]
    else:
        paths = [p for p in input_dir.iterdir() if p.is_file()]

    imgs = [p for p in paths if p.suffix.lower() in IMAGE_EXTS]
    imgs.sort()
    return imgs


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Batch crack detector runner (no GUI)")
    ap.add_argument("--input-dir", default="img_input", help="Folder containing input images")
    ap.add_argument("--config", default=None, help="Path to YAML/JSON config")
    ap.add_argument("--recursive", action="store_true", help="Scan input folder recursively")
    ap.add_argument("--disable-geometry", action="store_true", help="Disable geometry filter")
    ap.add_argument("--debug-enabled", action="store_true", help="Enable debug in runtime")
    ap.add_argument("--prompt-mode", default="disabled", help="Prompt mode: disabled | one_pass | multi_pass")
    ap.add_argument("--preprocess-variants", default="base", help="Comma-separated variants, e.g. base,blur_boost")
    ap.add_argument("--gt-dir", default="tests/masks", help="Folder containing GT masks")
    ap.add_argument("--gt-ext", default=".png", help="GT mask extension, e.g. .png")
    ap.add_argument("--gt-suffix", default="", help="Optional suffix for GT names, e.g. _mask")
    ap.add_argument("--pred-key", default="final_mask", help="Which result.images key to use as prediction mask")
    ap.add_argument("--pred-thr", type=float, default=0.5, help="Threshold for pred mask binarization")
    ap.add_argument("--gt-thr", type=float, default=0.5, help="Threshold for GT mask binarization")
    return ap.parse_args()


def _load_grayscale(path: Path) -> np.ndarray:
    m = cv2.imread(str(path), cv2.IMREAD_GRAYSCALE)
    if m is None:
        raise FileNotFoundError(f"Cannot read image: {path}")
    return m


def _to_binary_mask(arr: np.ndarray, thr: float) -> np.ndarray:
    # arr can be float/bool/uint8; output bool
    if arr.dtype == np.bool_:
        return arr
    a = arr
    if a.ndim == 3:
        # if HWC, take first channel
        a = a[:, :, 0]
    if np.issubdtype(a.dtype, np.floating):
        a = np.nan_to_num(a, nan=0.0, posinf=0.0, neginf=0.0)
        # if in 0..1 range, thr applies directly; else assume 0..255
        mx = float(np.max(a)) if a.size else 0.0
        if mx <= 1.0:
            return a >= thr
        return a >= (thr * 255.0)
    # integer types (e.g. uint8 0..255)
    return a >= int(thr * 255.0) if (a.max() > 1) else (a >= int(thr))


def _resize_to(a: np.ndarray, h: int, w: int) -> np.ndarray:
    if a.shape[0] == h and a.shape[1] == w:
        return a
    return cv2.resize(a, (w, h), interpolation=cv2.INTER_NEAREST)


def _confusion(pred: np.ndarray, gt: np.ndarray) -> tuple[int, int, int, int]:
    # pred, gt are bool with same shape
    tp = int(np.logical_and(pred, gt).sum())
    fp = int(np.logical_and(pred, ~gt).sum())
    fn = int(np.logical_and(~pred, gt).sum())
    tn = int(np.logical_and(~pred, ~gt).sum())
    return tp, fp, fn, tn


def _safe_div(a: float, b: float) -> float:
    return float(a / b) if b != 0 else 0.0


def _metrics_from_conf(tp: int, fp: int, fn: int, tn: int) -> dict[str, float]:
    precision = _safe_div(tp, tp + fp)
    recall = _safe_div(tp, tp + fn)
    iou = _safe_div(tp, tp + fp + fn)
    dice = _safe_div(2 * tp, 2 * tp + fp + fn)
    f1 = _safe_div(2 * precision * recall, precision + recall) if (precision + recall) != 0 else 0.0
    return {
        "iou": iou,
        "dice": dice,
        "precision": precision,
        "recall": recall,
        "f1": f1,
    }


def _find_gt_path(img_path: Path, gt_dir: Path, gt_suffix: str, gt_ext: str) -> Path:
    # Map by stem: anh1.jpg -> gt_dir/anh1{suffix}{ext}
    return gt_dir / f"{img_path.stem}{gt_suffix}{gt_ext}"


def _write_excel(path: Path, rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "per_image"

    if not rows:
        wb.save(str(path))
        return

    preferred = [
        "image",
        "status",
        "iou",
        "dice",
        "precision",
        "recall",
        "f1",
        "infer_ms",
        "tp",
        "fp",
        "fn",
        "tn",
        "image_path",
        "gt_path",
    ]
    keys_set = set().union(*[r.keys() for r in rows])
    fieldnames = [k for k in preferred if k in keys_set] + sorted([k for k in keys_set if k not in preferred])

    ws.append(fieldnames)
    for r in rows:
        ws.append([r.get(k) for k in fieldnames])

    for col_idx, k in enumerate(fieldnames, start=1):
        max_len = max(len(str(k)), *(len(str(r.get(k, ""))) for r in rows))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 60)

    wb.save(str(path))


def main() -> None:
    args = parse_args()

    input_dir = Path(args.input_dir)
    gt_dir = Path(args.gt_dir) if args.gt_dir else None
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_root = Path(f"result_{ts}")
    out_root.mkdir(parents=True, exist_ok=True)

    images = iter_images(input_dir, recursive=bool(args.recursive))
    if not images:
        print(f"No images found in: {input_dir.resolve()}")
        print("Put images into img_input/ and re-run.")
        return

    pipeline = CrackDetectionPipeline(args.config)

    variants = [v.strip() for v in str(args.preprocess_variants).split(",") if v.strip()]
    if not variants:
        variants = ["base"]

    runtime = {
        "enable_geometry_filter": not bool(args.disable_geometry),
        "debug_enabled": bool(args.debug_enabled),
        "preprocess_variants": variants,
        "damage_prompt_mode": str(args.prompt_mode),
    }

    aggregated_rows: list[dict[str, Any]] = []

    for img_path in images:
        name = _safe_name(img_path)
        img_out_dir = out_root / name
        step_dir = img_out_dir / "step_img"
        step_dir.mkdir(parents=True, exist_ok=True)

        t0 = time.perf_counter()
        try:
            result = pipeline.run(str(img_path), runtime=runtime)
        except Exception as exc:
            infer_ms = (time.perf_counter() - t0) * 1000.0
            err_path = img_out_dir / "error.txt"
            err_path.parent.mkdir(parents=True, exist_ok=True)
            err_path.write_text(str(exc), encoding="utf-8")
            print(f"FAILED: {img_path.name}: {exc}")

            aggregated_rows.append(
                {
                    "image": img_path.name,
                    "image_path": str(img_path.resolve()),
                    "infer_ms": infer_ms,
                    "status": "failed",
                    "error": str(exc),
                }
            )
            continue
        infer_ms = (time.perf_counter() - t0) * 1000.0

        # save step images
        stage_rows: list[dict[str, Any]] = []
        stages = _build_stage_list(result.images or {})
        for i, (k, stage_name) in enumerate(stages, start=1):
            v = (result.images or {}).get(k)
            if v is None or not isinstance(v, np.ndarray):
                continue
            base_name = f"{i:02d}_{stage_name}"
            out_path = _save_step_image(step_dir, base_name, v)
            st = _image_stats(v)
            stage_rows.append(
                {
                    "stage_index": int(i),
                    "stage_key": str(k),
                    "stage_name": str(stage_name),
                    "filename": str(out_path.name),
                    **st,
                }
            )
        _write_stages_csv(img_out_dir / "stages.csv", stage_rows)

        # save per-image metrics
        metrics_row: dict[str, Any] = {
            "image": img_path.name,
            "image_path": str(img_path.resolve()),
            "infer_ms": infer_ms,
        }

        # GT-based pixel metrics
        if gt_dir is not None:
            gt_path = _find_gt_path(img_path, gt_dir, str(args.gt_suffix), str(args.gt_ext))
            if not gt_path.exists():
                metrics_row.update({"status": "missing_gt", "gt_path": str(gt_path)})
            else:
                gt = _load_grayscale(gt_path)
                gt_bin = _to_binary_mask(gt, float(args.gt_thr))

                pred = None
                imgs = result.images or {}
                if str(args.pred_key) in imgs and isinstance(imgs[str(args.pred_key)], np.ndarray):
                    pred = imgs[str(args.pred_key)]
                else:
                    for k in ("final_mask", "geometry_kept_mask_viz", "sam_raw_mask_viz"):
                        if k in imgs and isinstance(imgs[k], np.ndarray):
                            pred = imgs[k]
                            break

                if pred is None:
                    metrics_row.update({"status": "missing_pred"})
                else:
                    pred2 = pred
                    if pred2.ndim == 3 and pred2.shape[2] == 1:
                        pred2 = pred2[:, :, 0]
                    pred2 = _resize_to(pred2, gt.shape[0], gt.shape[1])

                    pred_bin = _to_binary_mask(pred2, float(args.pred_thr))

                    tp, fp, fn, tn = _confusion(pred_bin, gt_bin)
                    m = _metrics_from_conf(tp, fp, fn, tn)

                    metrics_row.update(
                        {
                            "status": "ok",
                            "gt_path": str(gt_path.resolve()),
                            "tp": tp,
                            "fp": fp,
                            "fn": fn,
                            "tn": tn,
                            **m,
                        }
                    )
        else:
            metrics_row.update({"status": "no_gt_dir"})

        for k, v in (result.metrics or {}).items():
            if isinstance(v, (dict, list)):
                val: Any = json.dumps(v, ensure_ascii=False)
            else:
                val = v
            if k in metrics_row:
                metrics_row[f"model_{k}"] = val
            else:
                metrics_row[k] = val
        metrics_row["warnings"] = json.dumps(result.warnings or [], ensure_ascii=False)
        _write_metrics_csv(img_out_dir / "result.csv", metrics_row)

        _write_regions_csv(img_out_dir / "regions.csv", list(result.regions or []))

        aggregated_rows.append(metrics_row)

    _write_aggregated_csv(out_root / "result.csv", aggregated_rows)
    _write_excel(out_root / "result.xlsx", aggregated_rows)
    print(f"Excel written: {(out_root / 'result.xlsx').resolve()}")

    print(f"Done. Output: {out_root.resolve()}")


if __name__ == "__main__":
    main()
